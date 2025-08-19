import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Any
import json

class ExportService:
    
    @staticmethod
    def export_to_pdf(data: List[Dict[str, Any]], title: str, filename: str = None) -> BytesIO:
        buffer = BytesIO()
        
        if filename is None:
            filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1
        )
        
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 20))
        
        if data:
            headers = list(data[0].keys())
            table_data = [headers]
            
            for row in data:
                table_data.append([str(row.get(header, '')) for header in headers])
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_to_csv(data: List[Dict[str, Any]], filename: str = None) -> BytesIO:
        buffer = BytesIO()
        
        if filename is None:
            filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        if data:
            df = pd.DataFrame(data)
            df.to_csv(buffer, index=False)
        
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_to_excel(data: List[Dict[str, Any]], title: str, filename: str = None) -> BytesIO:
        buffer = BytesIO()
        
        if filename is None:
            filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        
        if data:
            headers = list(data[0].keys())
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Write data
            for row, data_row in enumerate(data, 2):
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=str(data_row.get(header, '')))
                    cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_data(data: List[Dict[str, Any]], format_type: str, title: str) -> tuple[BytesIO, str, str]:
        if format_type.lower() == "pdf":
            buffer = ExportService.export_to_pdf(data, title)
            content_type = "application/pdf"
            filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        elif format_type.lower() == "csv":
            buffer = ExportService.export_to_csv(data)
            content_type = "text/csv"
            filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        elif format_type.lower() == "excel":
            buffer = ExportService.export_to_excel(data, title)
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        else:
            buffer = ExportService.export_to_excel(data, title)
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return buffer, content_type, filename
